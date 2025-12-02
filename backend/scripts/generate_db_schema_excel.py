"""Script para gerar arquivo Excel com estrutura do banco de dados"""

import sys
from pathlib import Path

# Adiciona o diretório raiz ao path
sys.path.insert(0, str(Path(__file__).parent.parent))

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Instalando openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# Estrutura das tabelas baseada nos modelos
TABLES = [
    {
        "name": "users",
        "description": "Tabela de usuários do sistema",
        "columns": [
            {"name": "id", "type": "BigInteger", "nullable": False, "primary_key": True, "description": "ID único do usuário"},
            {"name": "email", "type": "String(255)", "nullable": False, "unique": True, "description": "Email do usuário (único)"},
            {"name": "password_hash", "type": "String(255)", "nullable": False, "description": "Hash da senha (bcrypt)"},
            {"name": "role", "type": "Enum(user_role)", "nullable": False, "description": "Papel: buyer, seller, service_provider"},
            {"name": "nickname", "type": "String(100)", "nullable": True, "description": "Apelido do usuário"},
            {"name": "email_verificado", "type": "Boolean", "nullable": False, "default": "False", "description": "Se o email foi verificado"},
            {"name": "created_at", "type": "DateTime", "nullable": False, "description": "Data de criação"},
            {"name": "updated_at", "type": "DateTime", "nullable": False, "description": "Data de atualização"},
        ]
    },
    {
        "name": "companies",
        "description": "Tabela de empresas/vendedores",
        "columns": [
            {"name": "id", "type": "BigInteger", "nullable": False, "primary_key": True, "description": "ID único da empresa"},
            {"name": "user_id", "type": "BigInteger", "nullable": False, "foreign_key": "users.id", "unique": True, "description": "FK para users"},
            {"name": "nome_propriedade", "type": "String(255)", "nullable": False, "description": "Nome da propriedade/empresa"},
            {"name": "inicio_atividades", "type": "Date", "nullable": True, "description": "Data de início das atividades"},
            {"name": "ramo_atividade", "type": "String(255)", "nullable": True, "description": "Ramo de atividade"},
            {"name": "cnaes", "type": "String(255)", "nullable": True, "description": "Códigos CNAE"},
            {"name": "cnpj_cpf", "type": "String(20)", "nullable": False, "description": "CNPJ ou CPF da empresa"},
            {"name": "insc_est_identidade", "type": "String(50)", "nullable": True, "description": "Inscrição estadual ou identidade"},
            {"name": "endereco", "type": "String(255)", "nullable": False, "description": "Endereço completo"},
            {"name": "bairro", "type": "String(100)", "nullable": True, "description": "Bairro"},
            {"name": "cep", "type": "String(12)", "nullable": False, "description": "CEP"},
            {"name": "cidade", "type": "String(100)", "nullable": False, "description": "Cidade"},
            {"name": "estado", "type": "String(2)", "nullable": False, "description": "Estado (UF)"},
            {"name": "email", "type": "String(255)", "nullable": False, "description": "Email de contato"},
            {"name": "created_at", "type": "DateTime", "nullable": False, "description": "Data de criação"},
            {"name": "updated_at", "type": "DateTime", "nullable": False, "description": "Data de atualização"},
        ]
    },
    {
        "name": "company_activities",
        "description": "Atividades das empresas (categoria, grupo, item)",
        "columns": [
            {"name": "id", "type": "BigInteger", "nullable": False, "primary_key": True, "description": "ID único"},
            {"name": "company_id", "type": "BigInteger", "nullable": False, "foreign_key": "companies.id", "description": "FK para companies"},
            {"name": "category_id", "type": "BigInteger", "nullable": False, "foreign_key": "activity_category.id", "description": "FK para activity_category"},
            {"name": "group_id", "type": "BigInteger", "nullable": True, "foreign_key": "activity_group.id", "description": "FK para activity_group"},
            {"name": "item_id", "type": "BigInteger", "nullable": True, "foreign_key": "activity_item.id", "description": "FK para activity_item"},
        ]
    },
    {
        "name": "buyer_profiles",
        "description": "Perfil de compradores",
        "columns": [
            {"name": "id", "type": "BigInteger", "nullable": False, "primary_key": True, "description": "ID único"},
            {"name": "user_id", "type": "BigInteger", "nullable": False, "foreign_key": "users.id", "unique": True, "description": "FK para users"},
            {"name": "nome_completo", "type": "String(255)", "nullable": False, "description": "Nome completo"},
            {"name": "data_nascimento", "type": "Date", "nullable": True, "description": "Data de nascimento"},
            {"name": "cpf", "type": "String(14)", "nullable": True, "unique": True, "description": "CPF (único)"},
            {"name": "identidade", "type": "String(20)", "nullable": True, "description": "RG"},
            {"name": "estado_civil", "type": "String(20)", "nullable": True, "description": "Estado civil"},
            {"name": "naturalidade", "type": "String(100)", "nullable": True, "description": "Naturalidade"},
            {"name": "endereco", "type": "String(255)", "nullable": False, "description": "Endereço completo"},
            {"name": "bairro", "type": "String(100)", "nullable": True, "description": "Bairro"},
            {"name": "cep", "type": "String(12)", "nullable": False, "description": "CEP"},
            {"name": "cidade", "type": "String(100)", "nullable": False, "description": "Cidade"},
            {"name": "estado", "type": "String(2)", "nullable": False, "description": "Estado (UF)"},
            {"name": "created_at", "type": "DateTime", "nullable": False, "description": "Data de criação"},
            {"name": "updated_at", "type": "DateTime", "nullable": False, "description": "Data de atualização"},
        ]
    },
    {
        "name": "service_providers",
        "description": "Prestadores de serviço",
        "columns": [
            {"name": "id", "type": "BigInteger", "nullable": False, "primary_key": True, "description": "ID único"},
            {"name": "user_id", "type": "BigInteger", "nullable": False, "foreign_key": "users.id", "unique": True, "description": "FK para users"},
            {"name": "nome_servico", "type": "String(255)", "nullable": False, "description": "Nome do serviço"},
            {"name": "descricao", "type": "String(500)", "nullable": True, "description": "Descrição do serviço"},
            {"name": "telefone", "type": "String(30)", "nullable": True, "description": "Telefone de contato"},
            {"name": "email_contato", "type": "String(255)", "nullable": False, "description": "Email de contato"},
            {"name": "cidade", "type": "String(100)", "nullable": False, "description": "Cidade"},
            {"name": "estado", "type": "String(2)", "nullable": False, "description": "Estado (UF)"},
            {"name": "tipo_servico", "type": "String(100)", "nullable": True, "description": "Tipo de serviço"},
            {"name": "endereco", "type": "String(255)", "nullable": True, "description": "Endereço"},
            {"name": "bairro", "type": "String(100)", "nullable": True, "description": "Bairro"},
            {"name": "cep", "type": "String(12)", "nullable": True, "description": "CEP"},
            {"name": "cnpj_cpf", "type": "String(20)", "nullable": True, "description": "CNPJ ou CPF"},
            {"name": "insc_est_identidade", "type": "String(50)", "nullable": True, "description": "Inscrição estadual ou identidade"},
            {"name": "created_at", "type": "DateTime", "nullable": False, "description": "Data de criação"},
            {"name": "updated_at", "type": "DateTime", "nullable": False, "description": "Data de atualização"},
        ]
    },
    {
        "name": "activity_category",
        "description": "Categorias de atividades (Agricultura, Pecuária, etc.)",
        "columns": [
            {"name": "id", "type": "BigInteger", "nullable": False, "primary_key": True, "description": "ID único"},
            {"name": "name", "type": "String(100)", "nullable": False, "unique": True, "description": "Nome da categoria"},
        ]
    },
    {
        "name": "activity_group",
        "description": "Grupos dentro de cada categoria",
        "columns": [
            {"name": "id", "type": "BigInteger", "nullable": False, "primary_key": True, "description": "ID único"},
            {"name": "category_id", "type": "BigInteger", "nullable": False, "foreign_key": "activity_category.id", "description": "FK para activity_category"},
            {"name": "name", "type": "String(100)", "nullable": False, "description": "Nome do grupo"},
        ]
    },
    {
        "name": "activity_item",
        "description": "Itens específicos dentro de cada grupo",
        "columns": [
            {"name": "id", "type": "BigInteger", "nullable": False, "primary_key": True, "description": "ID único"},
            {"name": "group_id", "type": "BigInteger", "nullable": False, "foreign_key": "activity_group.id", "description": "FK para activity_group"},
            {"name": "name", "type": "String(100)", "nullable": False, "description": "Nome do item"},
        ]
    },
    {
        "name": "quotations",
        "description": "Cotações/ofertas de produtos e serviços",
        "columns": [
            {"name": "id", "type": "BigInteger", "nullable": False, "primary_key": True, "description": "ID único"},
            {"name": "seller_id", "type": "BigInteger", "nullable": False, "foreign_key": "users.id", "description": "FK para users (vendedor)"},
            {"name": "seller_type", "type": "String(50)", "nullable": False, "description": "Tipo: 'company' ou 'service_provider'"},
            {"name": "title", "type": "String(255)", "nullable": False, "description": "Título da cotação"},
            {"name": "description", "type": "Text", "nullable": True, "description": "Descrição detalhada"},
            {"name": "category", "type": "Enum(quotation_category)", "nullable": False, "description": "agriculture, livestock, service, both"},
            {"name": "product_type", "type": "String(100)", "nullable": True, "description": "Tipo de produto (ex: 'boi', 'defensivo')"},
            {"name": "location_city", "type": "String(100)", "nullable": True, "description": "Cidade"},
            {"name": "location_state", "type": "String(2)", "nullable": True, "description": "Estado (UF)"},
            {"name": "price", "type": "Float", "nullable": True, "description": "Preço unitário"},
            {"name": "quantity", "type": "Float", "nullable": True, "description": "Quantidade disponível"},
            {"name": "unit", "type": "String(50)", "nullable": True, "description": "Unidade (kg, unidade, lote, hectare)"},
            {"name": "status", "type": "Enum(quotation_status)", "nullable": False, "default": "active", "description": "active, reserved, sold, expired, cancelled"},
            {"name": "expires_at", "type": "DateTime", "nullable": True, "description": "Data de expiração"},
            {"name": "image_url", "type": "String(500)", "nullable": True, "description": "URL da imagem principal"},
            {"name": "images", "type": "Text", "nullable": True, "description": "JSON array de URLs de imagens"},
            {"name": "free_shipping", "type": "Boolean", "nullable": False, "default": "False", "description": "Frete grátis"},
            {"name": "discount_percentage", "type": "Integer", "nullable": True, "description": "Percentual de desconto"},
            {"name": "installments", "type": "Integer", "nullable": True, "description": "Número de parcelas"},
            {"name": "stock", "type": "Integer", "nullable": True, "description": "Estoque disponível"},
            {"name": "embedding", "type": "Text", "nullable": True, "description": "JSON array do embedding (para IA matching)"},
            {"name": "created_at", "type": "DateTime", "nullable": False, "description": "Data de criação"},
            {"name": "updated_at", "type": "DateTime", "nullable": False, "description": "Data de atualização"},
        ]
    },
    {
        "name": "matches",
        "description": "Matches (Deu Agro) - correspondências entre compradores e cotações",
        "columns": [
            {"name": "id", "type": "BigInteger", "nullable": False, "primary_key": True, "description": "ID único"},
            {"name": "quotation_id", "type": "BigInteger", "nullable": False, "foreign_key": "quotations.id", "description": "FK para quotations"},
            {"name": "buyer_id", "type": "BigInteger", "nullable": False, "foreign_key": "users.id", "description": "FK para users (comprador)"},
            {"name": "score", "type": "Float", "nullable": True, "description": "Score de relevância calculado pela IA (0-100)"},
            {"name": "status", "type": "Enum(match_status)", "nullable": False, "default": "pending", "description": "pending, accepted, rejected, completed, cancelled"},
            {"name": "created_at", "type": "DateTime", "nullable": False, "description": "Data de criação"},
            {"name": "updated_at", "type": "DateTime", "nullable": False, "description": "Data de atualização"},
        ]
    },
    {
        "name": "email_verification_tokens",
        "description": "Tokens para verificação de email",
        "columns": [
            {"name": "id", "type": "BigInteger", "nullable": False, "primary_key": True, "description": "ID único"},
            {"name": "user_id", "type": "BigInteger", "nullable": False, "foreign_key": "users.id", "unique": True, "description": "FK para users"},
            {"name": "token", "type": "String(255)", "nullable": False, "unique": True, "description": "Token único de verificação"},
            {"name": "expires_at", "type": "DateTime", "nullable": False, "description": "Data de expiração do token"},
            {"name": "created_at", "type": "DateTime", "nullable": False, "description": "Data de criação"},
        ]
    },
    {
        "name": "password_reset_tokens",
        "description": "Tokens para recuperação de senha",
        "columns": [
            {"name": "id", "type": "BigInteger", "nullable": False, "primary_key": True, "description": "ID único"},
            {"name": "user_id", "type": "BigInteger", "nullable": False, "foreign_key": "users.id", "unique": True, "description": "FK para users"},
            {"name": "token", "type": "String(255)", "nullable": False, "unique": True, "description": "Token único de recuperação"},
            {"name": "expires_at", "type": "DateTime", "nullable": False, "description": "Data de expiração do token"},
            {"name": "used", "type": "Boolean", "nullable": False, "default": "False", "description": "Se o token já foi usado"},
            {"name": "created_at", "type": "DateTime", "nullable": False, "description": "Data de criação"},
        ]
    },
]


def create_excel():
    """Cria arquivo Excel com estrutura do banco de dados"""
    wb = openpyxl.Workbook()
    
    # Remove sheet padrão
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    title_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Cria índice
    index_sheet = wb.create_sheet("Índice", 0)
    index_sheet.append(["Tabela", "Descrição"])
    index_sheet["A1"].fill = title_fill
    index_sheet["A1"].font = title_font
    index_sheet["B1"].fill = title_fill
    index_sheet["B1"].font = title_font
    index_sheet["A1"].border = border
    index_sheet["B1"].border = border
    
    row = 2
    for table in TABLES:
        index_sheet.append([table["name"], table["description"]])
        index_sheet[f"A{row}"].border = border
        index_sheet[f"B{row}"].border = border
        row += 1
    
    # Ajusta largura das colunas do índice
    index_sheet.column_dimensions["A"].width = 30
    index_sheet.column_dimensions["B"].width = 60
    
    # Cria uma aba para cada tabela
    for table in TABLES:
        sheet = wb.create_sheet(table["name"])
        
        # Título
        sheet.merge_cells("A1:F1")
        title_cell = sheet["A1"]
        title_cell.value = f"Tabela: {table['name']}"
        title_cell.fill = title_fill
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.border = border
        
        # Descrição
        sheet.merge_cells("A2:F2")
        desc_cell = sheet["A2"]
        desc_cell.value = table["description"]
        desc_cell.alignment = Alignment(horizontal="center", vertical="center")
        desc_cell.border = border
        
        # Cabeçalhos
        headers = ["Coluna", "Tipo", "Nullable", "Chave", "Default", "Descrição"]
        sheet.append(headers)
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(row=3, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Dados das colunas
        for col_data in table["columns"]:
            row_data = [
                col_data["name"],
                col_data["type"],
                "Não" if not col_data.get("nullable", True) else "Sim",
                "",
                col_data.get("default", ""),
                col_data.get("description", "")
            ]
            
            # Adiciona informações de chave
            if col_data.get("primary_key"):
                row_data[3] = "PK"
            if col_data.get("foreign_key"):
                row_data[3] = f"FK: {col_data['foreign_key']}"
            if col_data.get("unique"):
                if row_data[3]:
                    row_data[3] += ", UNIQUE"
                else:
                    row_data[3] = "UNIQUE"
            
            sheet.append(row_data)
            
            # Aplica bordas
            for col in range(1, 7):
                cell = sheet.cell(row=sheet.max_row, column=col)
                cell.border = border
        
        # Ajusta largura das colunas
        sheet.column_dimensions["A"].width = 25
        sheet.column_dimensions["B"].width = 25
        sheet.column_dimensions["C"].width = 10
        sheet.column_dimensions["D"].width = 20
        sheet.column_dimensions["E"].width = 15
        sheet.column_dimensions["F"].width = 50
        
        # Congela primeira linha
        sheet.freeze_panes = "A4"
    
    # Salva arquivo
    output_path = Path(__file__).parent.parent.parent / "ESTRUTURA_BANCO_DADOS.xlsx"
    wb.save(output_path)
    print(f"✅ Arquivo Excel criado com sucesso: {output_path}")
    return output_path


if __name__ == "__main__":
    try:
        output_path = create_excel()
        print(f"\n📊 Arquivo gerado: {output_path}")
        print(f"📋 Total de tabelas: {len(TABLES)}")
    except Exception as e:
        print(f"❌ Erro ao gerar Excel: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

